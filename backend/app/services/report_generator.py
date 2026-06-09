"""
Servicio de generación de reportes en múltiples formatos: PDF, HTML, Excel.
Exporta datos de incidentes, finanzas y analytics.
"""
import io
from datetime import datetime
from typing import List
import pytz
from sqlalchemy.orm import Session
from sqlalchemy import func

# Importar generadores
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from jinja2 import Environment, BaseLoader

from app.models.incident import Incident
from app.models.payment import Payment
from app.models.user import User
from app.models.enums import IncidentStatus, PaymentStatus


BOL_TZ = pytz.timezone('America/La_Paz')


class ReportGenerator:
    """Generador de reportes en PDF, HTML y Excel."""

    @staticmethod
    def generate_incidents_report(
        db: Session,
        tenant_id: int,
        start_date: datetime,
        end_date: datetime,
        format: str = "pdf",  # pdf, html, excel
    ) -> bytes:
        """
        Genera reporte de incidentes entre fechas.
        
        Args:
            format: "pdf", "html" o "excel"
        
        Returns:
            bytes: Contenido del archivo
        """
        incidents = db.query(Incident).filter(
            Incident.tenant_id == tenant_id,
            Incident.created_at >= start_date,
            Incident.created_at <= end_date,
        ).order_by(Incident.created_at.desc()).all()

        data = [
            {
                "id": inc.id,
                "user": inc.user.full_name if inc.user else "N/A",
                "vehicle": f"{inc.vehicle.brand} {inc.vehicle.model}" if inc.vehicle else "N/A",
                "type": inc.incident_type.value,
                "status": inc.status.value,
                "priority": inc.priority.value,
                "created_at": inc.created_at.astimezone(BOL_TZ).strftime("%Y-%m-%d %H:%M"),
                "cost": inc.final_cost or 0,
                "cancellation_fee": inc.cancellation_fee or 0,
            }
            for inc in incidents
        ]

        if format == "pdf":
            return ReportGenerator._generate_pdf_incidents(data, start_date, end_date)
        elif format == "html":
            return ReportGenerator._generate_html_incidents(data, start_date, end_date)
        elif format == "excel":
            return ReportGenerator._generate_excel_incidents(data, start_date, end_date)

    @staticmethod
    def generate_financial_report(
        db: Session,
        tenant_id: int,
        start_date: datetime,
        end_date: datetime,
        format: str = "pdf",
    ) -> bytes:
        """Genera reporte financiero (pagos, comisiones, etc.)."""
        
        payments = db.query(Payment).filter(
            Payment.tenant_id == tenant_id,
            Payment.created_at >= start_date,
            Payment.created_at <= end_date,
        ).all()

        total_amount = sum(p.amount for p in payments)
        total_commission = sum(p.commission_amount for p in payments)
        total_completed = sum(p.amount for p in payments if p.payment_status == PaymentStatus.COMPLETED)
        total_pending = sum(p.amount for p in payments if p.payment_status == PaymentStatus.PENDING)
        total_cancellation_fees = sum(p.cancellation_fee for p in payments if p.cancellation_fee)

        data = {
            "period": f"{start_date.strftime('%Y-%m-%d')} a {end_date.strftime('%Y-%m-%d')}",
            "total_payments": len(payments),
            "total_amount": round(total_amount, 2),
            "total_commission": round(total_commission, 2),
            "total_completed": round(total_completed, 2),
            "total_pending": round(total_pending, 2),
            "total_cancellation_fees": round(total_cancellation_fees, 2),
            "net_amount": round(total_completed - total_commission, 2),
            "details": [
                {
                    "id": p.id,
                    "incident_id": p.incident_id,
                    "amount": p.amount,
                    "commission": p.commission_amount,
                    "status": p.payment_status.value,
                    "method": p.payment_method.value if p.payment_method else "N/A",
                    "created_at": p.created_at.astimezone(BOL_TZ).strftime("%Y-%m-%d %H:%M"),
                }
                for p in payments
            ]
        }

        if format == "pdf":
            return ReportGenerator._generate_pdf_financial(data)
        elif format == "html":
            return ReportGenerator._generate_html_financial(data)
        elif format == "excel":
            return ReportGenerator._generate_excel_financial(data)
    # ========== GENERADORES PDF ==========
    @staticmethod
    def _generate_pdf_incidents(data: list, start_date, end_date) -> bytes:
        """Genera PDF de incidentes."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=12,
            alignment=TA_CENTER,
        )

        cell_style = ParagraphStyle(
            'CellText',
            parent=styles['Normal'],
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor('#2c3e50'),
            alignment=TA_CENTER,
        )
        
        header_style = ParagraphStyle(
            'HeaderText',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # Título
        title = Paragraph(
            f"REPORTE DE INCIDENTES<br/>{start_date.strftime('%Y-%m-%d')} a {end_date.strftime('%Y-%m-%d')}",
            title_style
        )
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))

        # Tabla
        if data:
            headers = [
                Paragraph("ID", header_style),
                Paragraph("Usuario", header_style),
                Paragraph("Vehículo", header_style),
                Paragraph("Tipo", header_style),
                Paragraph("Estado", header_style),
                Paragraph("Prioridad", header_style),
                Paragraph("Fecha", header_style),
                Paragraph("Costo", header_style),
                Paragraph("Penalización", header_style)
            ]
            
            table_data = [headers]
            for row in data:
                table_data.append([
                    Paragraph(str(row["id"]), cell_style),
                    Paragraph(row["user"], cell_style),
                    Paragraph(row["vehicle"], cell_style),
                    Paragraph(row["type"], cell_style),
                    Paragraph(row["status"], cell_style),
                    Paragraph(row["priority"], cell_style),
                    Paragraph(row["created_at"], cell_style),
                    Paragraph(f"BOB {row['cost']}", cell_style),
                    Paragraph(f"BOB {row['cancellation_fee']}" if row['cancellation_fee'] else "-", cell_style),
                ])

            # Anchos proporcionales sumando 540 puntos (7.5 pulgadas de ancho imprimible)
            col_widths = [30, 95, 80, 50, 60, 50, 80, 50, 45]
            
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
            ]))
            elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _generate_pdf_financial(data: dict) -> bytes:
        """Genera PDF de reporte financiero."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2e7d32'),
            spaceAfter=12,
            alignment=TA_CENTER,
        )

        title = Paragraph(f"REPORTE FINANCIERO<br/>{data['period']}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))

        # Resumen
        summary_data = [
            ["MÉTRICA", "VALOR"],
            ["Total de Pagos", str(data["total_payments"])],
            ["Monto Total", f"BOB {data['total_amount']}"],
            ["Comisión Total", f"BOB {data['total_commission']}"],
            ["Pagos Completados", f"BOB {data['total_completed']}"],
            ["Pagos Pendientes", f"BOB {data['total_pending']}"],
            ["Penalizaciones", f"BOB {data['total_cancellation_fees']}"],
            ["Monto Neto", f"BOB {data['net_amount']}"],
        ]

        summary_table = Table(summary_data, colWidths=[220, 150])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e7d32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ]))
        elements.append(summary_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # ========== GENERADORES HTML ==========
    @staticmethod
    def _generate_html_incidents(data: list, start_date, end_date) -> bytes:
        """Genera HTML de incidentes."""
        html_template = """
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <title>Reporte de Incidentes</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; color: #333; }
                h1 { color: #1f77b4; text-align: center; }
                table { width: 100%; border-collapse: collapse; margin-top: 20px; }
                th { background-color: #1f77b4; color: white; padding: 10px; text-align: left; }
                td { padding: 8px; border-bottom: 1px solid #ddd; }
                tr:hover { background-color: #f5f5f5; }
                .summary { background-color: #f9f9f9; padding: 15px; margin: 20px 0; border-left: 4px solid #1f77b4; }
            </style>
        </head>
        <body>
            <h1>Reporte de Incidentes</h1>
            <p style="text-align: center;">{{ start_date }} a {{ end_date }}</p>
            <div class="summary">
                <p><strong>Total de Incidentes:</strong> {{ total_incidents }}</p>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Usuario</th>
                        <th>Vehículo</th>
                        <th>Tipo</th>
                        <th>Estado</th>
                        <th>Prioridad</th>
                        <th>Fecha</th>
                        <th>Costo (BOB)</th>
                        <th>Penalización (BOB)</th>
                    </tr>
                </thead>
                <tbody>
                    {% for row in data %}
                    <tr>
                        <td>{{ row.id }}</td>
                        <td>{{ row.user }}</td>
                        <td>{{ row.vehicle }}</td>
                        <td>{{ row.type }}</td>
                        <td>{{ row.status }}</td>
                        <td>{{ row.priority }}</td>
                        <td>{{ row.created_at }}</td>
                        <td>{{ row.cost }}</td>
                        <td>{{ row.cancellation_fee if row.cancellation_fee else '-' }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </body>
        </html>
        """
        
        env = Environment(loader=BaseLoader())
        template = env.from_string(html_template)
        html_content = template.render(
            start_date=start_date.strftime('%Y-%m-%d'),
            end_date=end_date.strftime('%Y-%m-%d'),
            total_incidents=len(data),
            data=data,
        )
        return html_content.encode('utf-8')

    @staticmethod
    def _generate_html_financial(data: dict) -> bytes:
        """Genera HTML de reporte financiero."""
        html_template = """
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <title>Reporte Financiero</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; color: #333; }
                h1 { color: #27ae60; text-align: center; }
                .summary { background-color: #f0fdf4; padding: 20px; border-radius: 8px; margin: 20px 0; }
                .metric { display: inline-block; width: 45%; padding: 10px; margin: 5px; background: white; border: 1px solid #ddd; border-radius: 5px; }
                table { width: 100%; border-collapse: collapse; margin-top: 20px; }
                th { background-color: #27ae60; color: white; padding: 10px; text-align: left; }
                td { padding: 8px; border-bottom: 1px solid #ddd; }
            </style>
        </head>
        <body>
            <h1>Reporte Financiero</h1>
            <p style="text-align: center;">Período: {{ period }}</p>
            <div class="summary">
                <div class="metric"><strong>Total de Pagos:</strong> {{ total_payments }}</div>
                <div class="metric"><strong>Monto Total:</strong> BOB {{ total_amount }}</div>
                <div class="metric"><strong>Comisión:</strong> BOB {{ total_commission }}</div>
                <div class="metric"><strong>Completados:</strong> BOB {{ total_completed }}</div>
                <div class="metric"><strong>Pendientes:</strong> BOB {{ total_pending }}</div>
                <div class="metric"><strong>Penalizaciones:</strong> BOB {{ total_cancellation_fees }}</div>
                <div class="metric"><strong>Monto Neto:</strong> BOB {{ net_amount }}</div>
            </div>
        </body>
        </html>
        """
        
        env = Environment(loader=BaseLoader())
        template = env.from_string(html_template)
        html_content = template.render(**data)
        return html_content.encode('utf-8')

    # ========== GENERADORES EXCEL ==========
    @staticmethod
    def _generate_excel_incidents(data: list, start_date, end_date) -> bytes:
        """Genera Excel de incidentes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Incidentes"

        # Encabezado
        ws['A1'] = f"Reporte de Incidentes - {start_date.strftime('%Y-%m-%d')} a {end_date.strftime('%Y-%m-%d')}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
        ws.merge_cells('A1:I1')

        # Encabezados de columna
        headers = ["ID", "Usuario", "Vehículo", "Tipo", "Estado", "Prioridad", "Fecha", "Costo (BOB)", "Penalización (BOB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for row_idx, row_data in enumerate(data, 4):
            ws.cell(row=row_idx, column=1, value=row_data["id"])
            ws.cell(row=row_idx, column=2, value=row_data["user"])
            ws.cell(row=row_idx, column=3, value=row_data["vehicle"])
            ws.cell(row=row_idx, column=4, value=row_data["type"])
            ws.cell(row=row_idx, column=5, value=row_data["status"])
            ws.cell(row=row_idx, column=6, value=row_data["priority"])
            ws.cell(row=row_idx, column=7, value=row_data["created_at"])
            ws.cell(row=row_idx, column=8, value=row_data["cost"])
            ws.cell(row=row_idx, column=9, value=row_data["cancellation_fee"] if row_data["cancellation_fee"] else "-")

        # Ajustar ancho de columnas
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _generate_excel_financial(data: dict) -> bytes:
        """Genera Excel de reporte financiero."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Financiero"

        # Encabezado
        ws['A1'] = f"Reporte Financiero - {data['period']}"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
        ws.merge_cells('A1:B1')

        # Resumen
        row = 3
        summary_fields = [
            ("Total de Pagos", data["total_payments"]),
            ("Monto Total (BOB)", data["total_amount"]),
            ("Comisión Total (BOB)", data["total_commission"]),
            ("Pagos Completados (BOB)", data["total_completed"]),
            ("Pagos Pendientes (BOB)", data["total_pending"]),
            ("Penalizaciones (BOB)", data["total_cancellation_fees"]),
            ("Monto Neto (BOB)", data["net_amount"]),
        ]

        for label, value in summary_fields:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
